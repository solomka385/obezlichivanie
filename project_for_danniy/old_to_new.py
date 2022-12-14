import openpyxl as px


#book = px.open('name_new.xlsx')
#old_data = 'Fred'
#data = 'John'   # Что нужно вписать в ячейку
#data_type = 'name'  # Название столбца


def id_exist(old_data, book, data):
    booked = px.open(book)
    page = booked.active  # Берём 1 лист в excel
    flag = 0
    for row in range(2, page.max_row+1):
        if page.cell(row, 1).value == old_data:    # Если id уже есть
            flag = 1
            page.cell(row, 2, data)

    if flag == 0:   # Если id нет
        row = page.max_row + 1
        page.cell(row, 1, old_data)
        page.cell(row, 2, data)
    booked.save(book)


#id_exist('fefe',  'C://Users/alex1/Desktop/project_for_danniy/name_new.xlsx','dwdw')
